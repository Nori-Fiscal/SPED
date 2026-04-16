from __future__ import annotations

import io
import re
import traceback
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd
import streamlit as st

FORMATO_CONTABIL = '_-[$R$-pt-BR] * #,##0.00_-'
FORMATO_ZERO = '_-[$R$-pt-BR] * -_-'


@dataclass
class ResultadoProcessamento:
    sucesso: bool
    mensagens: list[str]
    caminho_saida: str | None = None
    bytes_arquivo: bytes | None = None
    nome_origem: str | None = None


def converter_para_numero(valor):
    if valor is None:
        return 0
    if isinstance(valor, (int, float)):
        return valor
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, str):
        valor_limpo = re.sub(r'[^\d.,-]', '', valor.strip())
        if not valor_limpo:
            return 0
        if ',' in valor_limpo and '.' in valor_limpo:
            if valor_limpo.rfind(',') > valor_limpo.rfind('.'):
                valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
            else:
                valor_limpo = valor_limpo.replace(',', '')
        else:
            valor_limpo = valor_limpo.replace(',', '.')
        if valor_limpo.count('.') > 1:
            partes = valor_limpo.split('.')
            valor_limpo = ''.join(partes[:-1]) + '.' + partes[-1]
        try:
            return float(valor_limpo)
        except ValueError:
            return 0
    return 0


def ultima_linha_com_dados(ws, col_idx: int = 2) -> int:
    for i in range(ws.max_row, 1, -1):
        if ws.cell(row=i, column=col_idx).value not in (None, ''):
            return i
    return ws.max_row


def ajustar_largura_colunas(ws):
    for coluna in ws.columns:
        max_len = 0
        letra = coluna[0].column_letter
        for cell in coluna[:300]:
            valor = '' if cell.value is None else str(cell.value)
            if len(valor) > max_len:
                max_len = len(valor)
        ws.column_dimensions[letra].width = min(max(max_len + 2, 12), 40)


def processar_aba_d100(ws, mensagens: list[str]):
    if ws.max_row < 2:
        mensagens.append('Aba D100 vazia ou sem dados suficientes.')
        return

    ws.delete_rows(1)
    ws.insert_rows(1)
    titulos = [
        'IND_OPER1', 'IND_OPER', 'IND_EMIT', 'COD_PART', 'COD_MOD', 'COD_SIT', 'SER', 'SUB', 'NUM_DOC', 'CHV_CTE',
        'DT_DOC2', 'DT_DOC', 'TP_CT-e', 'CHV_CTE_REF', 'VL_DOC', 'VL_DESC', 'IND_FRT', 'VL_SERV', 'VL_BC_ICMS', 'VL_ICMS',
        'VL_NT', 'COD_INF', 'COD_CTA', 'COD_MUN_ORIG', 'COD_MUN_DEST'
    ]
    for i, titulo in enumerate(titulos, start=1):
        ws.cell(row=1, column=i).value = titulo

    ws.delete_cols(1)
    ws.delete_cols(10)

    ult = ultima_linha_com_dados(ws, 2)
    colunas_converter = ['H', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
    colunas_formatar = ['M', 'N', 'O', 'P', 'Q', 'R', 'S']

    for col in colunas_converter:
        idx = openpyxl.utils.column_index_from_string(col)
        for row in range(2, ult + 1):
            ws.cell(row=row, column=idx).value = converter_para_numero(ws.cell(row=row, column=idx).value)

    for col in colunas_formatar:
        idx = openpyxl.utils.column_index_from_string(col)
        for row in range(2, ult + 1):
            cell = ws.cell(row=row, column=idx)
            cell.number_format = FORMATO_ZERO if cell.value == 0 else FORMATO_CONTABIL

    ajustar_largura_colunas(ws)
    ws.freeze_panes = 'A2'
    mensagens.append('Aba D100 processada com sucesso.')


def processar_aba_c100(ws, mensagens: list[str]):
    if ws.max_row < 2:
        mensagens.append('Aba C100 vazia ou sem dados suficientes.')
        return

    ws.delete_rows(1)
    ws.insert_rows(1)
    titulos = [
        'REG', 'IND_OPER', 'IND_EMIT', 'COD_PART', 'COD_MOD', 'COD_SIT', 'SER', 'NUM_DOC', 'CHV_NFE',
        'DT_DOC', 'DT_DOC', 'VL_DOC', 'IND_PGTO', 'VL_DESC', 'VL_ABAT_NT', 'VL_MERC', 'IND_FRT', 'VL_FRT', 'VL_SEG',
        'VL_OUT_DA', 'VL_BC_ICMS', 'VL_ICMS', 'VL_BC_ICMS_ST', 'VL_ICMS_ST', 'VL_IPI', 'VL_PIS', 'VL_COFINS',
        'VL_PIS_ST', 'VL_COFINS_ST'
    ]
    for i, titulo in enumerate(titulos, start=1):
        ws.cell(row=1, column=i).value = titulo

    ws.delete_cols(10)
    ult = ultima_linha_com_dados(ws, 2)

    colunas_converter = ['H', 'K', 'M', 'N', 'O', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']
    colunas_formatar = ['K', 'M', 'N', 'O', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']

    for col in colunas_converter:
        idx = openpyxl.utils.column_index_from_string(col)
        for row in range(2, ult + 1):
            ws.cell(row=row, column=idx).value = converter_para_numero(ws.cell(row=row, column=idx).value)

    idx_j = openpyxl.utils.column_index_from_string('J')
    for row in range(2, ult + 1):
        cell = ws.cell(row=row, column=idx_j)
        valor = cell.value
        if isinstance(valor, str) and len(valor) == 8 and valor.isdigit():
            try:
                data = datetime(int(valor[4:]), int(valor[2:4]), int(valor[:2]))
                cell.value = data
                cell.number_format = 'dd/mm/yyyy'
            except ValueError:
                pass

    for col in colunas_formatar:
        idx = openpyxl.utils.column_index_from_string(col)
        for row in range(2, ult + 1):
            cell = ws.cell(row=row, column=idx)
            cell.number_format = FORMATO_ZERO if cell.value == 0 else FORMATO_CONTABIL

    ajustar_largura_colunas(ws)
    ws.freeze_panes = 'A2'
    mensagens.append('Aba C100 processada com sucesso.')


def converter_txt_sped_para_excel_bytes(arquivo_txt_bytes: bytes) -> tuple[bytes, list[str]]:
    mensagens: list[str] = []
    registros_por_tipo: dict[str, list[list[str]]] = defaultdict(list)

    texto = arquivo_txt_bytes.decode('latin1', errors='ignore').splitlines()
    for linha in texto:
        linha = linha.strip()
        if linha.startswith('|'):
            partes = linha.split('|')
            if len(partes) > 2:
                tipo = partes[1].strip() or 'SEM_TIPO'
                registros_por_tipo[tipo].append(partes[1:-1])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        for tipo, registros in registros_por_tipo.items():
            if len(registros) > 0:
                max_campos = max(len(r) for r in registros)
                colunas = [f'Campo_{i}' for i in range(1, max_campos + 1)]
                df = pd.DataFrame(registros, columns=colunas)
                aba = tipo[:31] or 'SEM_TIPO'
                df.to_excel(writer, sheet_name=aba, index=False)
                mensagens.append(f'Aba {aba} criada com {len(df)} linhas.')

    output.seek(0)
    return output.getvalue(), mensagens


def processar_excel_sped_bytes(arquivo_excel_bytes: bytes, abas_selecionadas: Iterable[str]) -> tuple[bytes, list[str]]:
    mensagens: list[str] = []
    arquivo = io.BytesIO(arquivo_excel_bytes)
    wb = openpyxl.load_workbook(arquivo)

    if 'D100' in abas_selecionadas and 'D100' in wb.sheetnames:
        processar_aba_d100(wb['D100'], mensagens)
    elif 'D100' in abas_selecionadas:
        mensagens.append('Aba D100 não encontrada no arquivo.')

    if 'C100' in abas_selecionadas and 'C100' in wb.sheetnames:
        processar_aba_c100(wb['C100'], mensagens)
    elif 'C100' in abas_selecionadas:
        mensagens.append('Aba C100 não encontrada no arquivo.')

    saida = io.BytesIO()
    wb.save(saida)
    saida.seek(0)
    return saida.getvalue(), mensagens


def processar_fluxo_completo(arquivo, abas_selecionadas: Iterable[str]) -> ResultadoProcessamento:
    mensagens: list[str] = []
    try:
        nome = arquivo.name.lower()
        conteudo = arquivo.getvalue()

        if nome.endswith('.txt'):
            mensagens.append('Convertendo TXT do SPED para Excel...')
            excel_bytes, msgs_conv = converter_txt_sped_para_excel_bytes(conteudo)
            mensagens.extend(msgs_conv)
            mensagens.append('Aplicando tratamento nas abas selecionadas...')
            final_bytes, msgs_proc = processar_excel_sped_bytes(excel_bytes, abas_selecionadas)
            mensagens.extend(msgs_proc)
            nome_saida = f"{Path(arquivo.name).stem}_processado.xlsx"
        elif nome.endswith('.xlsx'):
            mensagens.append('Processando Excel enviado...')
            final_bytes, msgs_proc = processar_excel_sped_bytes(conteudo, abas_selecionadas)
            mensagens.extend(msgs_proc)
            nome_saida = f"{Path(arquivo.name).stem}_processado.xlsx"
        else:
            return ResultadoProcessamento(False, ['Envie arquivos .txt ou .xlsx.'], nome_origem=arquivo.name)

        return ResultadoProcessamento(True, mensagens, nome_saida, final_bytes, arquivo.name)
    except Exception:
        mensagens.append('Erro durante o processamento:')
        mensagens.append(traceback.format_exc())
        return ResultadoProcessamento(False, mensagens, nome_origem=arquivo.name)


def salvar_em_disco(resultado: ResultadoProcessamento, pasta_saida: str) -> str:
    if not resultado.bytes_arquivo or not resultado.caminho_saida:
        raise ValueError('Nenhum arquivo para salvar.')
    pasta = Path(pasta_saida)
    pasta.mkdir(parents=True, exist_ok=True)
    destino = pasta / resultado.caminho_saida
    destino.write_bytes(resultado.bytes_arquivo)
    return str(destino)


def gerar_zip_resultados(resultados: list[ResultadoProcessamento]) -> bytes:
    memoria = io.BytesIO()
    with zipfile.ZipFile(memoria, 'w', zipfile.ZIP_DEFLATED) as zf:
        for resultado in resultados:
            if resultado.sucesso and resultado.bytes_arquivo and resultado.caminho_saida:
                zf.writestr(resultado.caminho_saida, resultado.bytes_arquivo)
            elif resultado.nome_origem:
                log_nome = f"{Path(resultado.nome_origem).stem}_erro.txt"
                zf.writestr(log_nome, '\n'.join(resultado.mensagens))
    memoria.seek(0)
    return memoria.getvalue()


def resumir_resultados(resultados: list[ResultadoProcessamento]) -> pd.DataFrame:
    linhas = []
    for resultado in resultados:
        linhas.append(
            {
                'Arquivo de origem': resultado.nome_origem or '',
                'Status': 'Sucesso' if resultado.sucesso else 'Falha',
                'Arquivo gerado': resultado.caminho_saida or '',
                'Mensagens': ' | '.join(resultado.mensagens[:4]),
            }
        )
    return pd.DataFrame(linhas)


def interface_streamlit():
    st.set_page_config(page_title='SPED Unificado', page_icon='📊', layout='wide')
    st.title('📊 SPED Unificado')
    st.caption('Converte vários TXT do SPED em Excel e processa automaticamente as abas D100 e C100.')

    with st.sidebar:
        st.subheader('Configurações')
        abas = st.multiselect('Abas para processar', ['D100', 'C100'], default=['D100', 'C100'])
        salvar_disco = st.checkbox('Salvar também em pasta local do computador', value=False)
        pasta_saida = st.text_input('Pasta de saída', value=str(Path.home() / 'Downloads'))
        gerar_zip = st.checkbox('Gerar ZIP com todos os arquivos', value=True)

    arquivos = st.file_uploader(
        'Selecione um ou vários arquivos SPED (.txt ou .xlsx)',
        type=['txt', 'xlsx'],
        accept_multiple_files=True,
    )

    col1, col2 = st.columns([1, 1])
    with col1:
        processar = st.button('Processar arquivos', type='primary', use_container_width=True, disabled=not arquivos)
    with col2:
        st.markdown('**Fluxo:** vários TXT/XLSX → processamento individual → download por arquivo ou em ZIP.')

    if processar and arquivos:
        if not abas:
            st.warning('Selecione pelo menos uma aba.')
            st.stop()

        resultados: list[ResultadoProcessamento] = []
        barra = st.progress(0, text='Iniciando processamento...')

        for indice, arquivo in enumerate(arquivos, start=1):
            barra.progress((indice - 1) / len(arquivos), text=f'Processando {arquivo.name}...')
            resultado = processar_fluxo_completo(arquivo, abas)
            resultados.append(resultado)

            if salvar_disco and resultado.sucesso:
                try:
                    caminho = salvar_em_disco(resultado, pasta_saida)
                    resultado.mensagens.append(f'Arquivo salvo em disco: {caminho}')
                except Exception as e:
                    resultado.mensagens.append(f'Falha ao salvar em disco: {e}')

        barra.progress(1.0, text='Processamento concluído.')

        total = len(resultados)
        sucessos = sum(1 for r in resultados if r.sucesso)
        falhas = total - sucessos

        if sucessos:
            st.success(f'{sucessos} arquivo(s) processado(s) com sucesso.')
        if falhas:
            st.warning(f'{falhas} arquivo(s) tiveram falha. Veja o log abaixo.')

        st.dataframe(resumir_resultados(resultados), use_container_width=True, hide_index=True)

        with st.expander('Ver logs detalhados', expanded=False):
            for resultado in resultados:
                st.markdown(f"### {resultado.nome_origem}")
                for msg in resultado.mensagens:
                    if 'Traceback' in msg:
                        st.code(msg)
                    else:
                        st.write('- ', msg)

        if gerar_zip and resultados:
            zip_bytes = gerar_zip_resultados(resultados)
            st.download_button(
                label='Baixar ZIP com todos os resultados',
                data=zip_bytes,
                file_name='SPED_processados.zip',
                mime='application/zip',
                use_container_width=True,
            )

        if len(resultados) == 1 and resultados[0].sucesso:
            resultado = resultados[0]
            st.download_button(
                label='Baixar arquivo processado',
                data=resultado.bytes_arquivo,
                file_name=resultado.caminho_saida,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )
        elif len(resultados) > 1:
            st.markdown('### Downloads individuais')
            for resultado in resultados:
                if resultado.sucesso and resultado.bytes_arquivo and resultado.caminho_saida:
                    st.download_button(
                        label=f"Baixar {resultado.caminho_saida}",
                        data=resultado.bytes_arquivo,
                        file_name=resultado.caminho_saida,
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        key=f"download_{resultado.caminho_saida}",
                        use_container_width=True,
                    )


if __name__ == '__main__':
    interface_streamlit()
